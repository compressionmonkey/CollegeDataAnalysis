# Read in excel data
from openpyxl import utils
from openpyxl import workbook
from openpyxl import load_workbook
wb = load_workbook('Hiace2015.xlsx', data_only=True) #show real number instead of excel formulas
print wb.worksheets

startRow = 0
endRow = 0
startCol = 0
endCol = 0



for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        #print row
        for cell in row:
            if cell.value == 'SL #': # Sets a condition for finding SL to assign values to our range
                coord = cell.coordinate
                t = utils.coordinate_from_string(coord)
                startCol = utils.column_index_from_string(t[0])
                #startCol = t[0]
                startRow = t[1]+4
#
            cellValue = cell.value
            if type(cellValue) is unicode:
                if cellValue.strip() == 'Total':
                    coord = cell.coordinate
                    t = utils.coordinate_from_string(coord)
                    endCol = utils.column_index_from_string(t[0])+8
                    #startCol = t[0]
                    endRow = t[1]-1

                if cellValue.strip() == 'Date' and cellValue.strip() == 'KM Covered' and cellValue.strip() == 'Amount':
                    print cellValue
#
# #for :
#  #   print range
    for irow in range(startRow, endRow):  # row 11 to 31
        row = sheet[irow]
        for icell in range(startCol, endCol):  # cell C to M
            if cell.coordinate == "1":
                a = cell.coordinate
                a = a + 1

            cell = row[icell]
            # if cell.coordinate > 'C10' and cell.coordinate < 'D12':
            print "You are currently on sheet ", sheet.title, 'and cell number', cell.coordinate
            # print cell.coordinate
            print cell.value

#print ("Highest col",icell.get_highest_column())

Final_ws = wb.create_sheet()
Final_ws.title = "All Data"
#wb.save('Hiace2015New.xlsx')







# sheet = dict(wb.active) # activating is an attribute and not a method
# sheetnames = list(wb.get_sheet_names())
# print sheetnames
# sum(sheet['C11:M31'].isnull())
# #sheet = int(sheet)
# cells = sheet['C11:M31']
# for c1, c2,c3,c4,c5,c6,c7,c8,c9,c10,c11 in cells:
#     print "{0:8} {1:8} {2:8} {3:8} {4:8} {5:8} {6:8} {7:8} {8:8} {9:8} {10:8}".format(c1.value, c2.value, c3.value, c4.value, c5.value, c6.value, c7.value, c8.value,
#                                c9.value, c10.value, c11.value)
#
# for row in sheet.iter_rows(cells):
#     for cell in row:
#         cells.append(cell.value)
#     print("Number of values: {0}".format(len(values)))
# def get_all_data(sheet_in):
#       Final_copy = Workbook()
#       Final_copy = Final_copy.active()
#       Final_copy.title = "Hiace Details"
#       Final_copy = Final_copy["Hiace Details"]
#       for row in row_range:
#           for column in col_range:
#             print row, column
#
#       wb.copy_worksheet()
#       wb.save('Final_Copy_Hiace_2016.xlsx')


            # create csv file that has all details to report on
# print ("Rows: ", sheet.max_row) # for debugging purposes
# print ("Columns: ", sheet.max_column) # for debugging purposes
# last_data_point = ws.cell(row = sheet.max_row, column = sheet.max_column).coordinate
# print ("Last data point in current worksheet:", last_data_point) #for debugging purposes
#
# #import next file and add to master
# append_point = ws.cell(row = sheet.max_row + 1, column = 1).coordinate
# print ("Start new data at:", append_point)
# wb = load_workbook('second_file.xlsx')
# sheet2 = wb.get_sheet_by_name('Sheet1')
# start = ws.cell(coordinate='A2').coordinate
# print("New data start: ", start)
# end = ws.cell(row = sheet2.max_row, column = sheet2.max_column).coordinate
# print ("New data end: ", end)
#
# # write a value to selected cell
# #sheet[append_point] = 311
# #print (ws.cell(append_point).value)
#
# #save file
# wb.save('master_file.xlsx')



#
# import numpy as np
# def extract_data(filename):
#
#     #array to hold the labels and feature vectors.
#     labels = []
#     fvecs = []
#
#     #iterate over the rows, split the label from the features
#     #convert labels to integers and features to floats
#     for line in file(filename):
#         row = line.split(',')
#         labels.append(int(row[0]))
#         fvecs.append([float(x) for x in row[1:2]])

    #convert the array of float arrays into a numpy float matrix
    # fvecs_np = np.matrix(fvecs).astype(np.float32)
    #
    # #convert the array of int labels into a numpy array
    # labels_np = np.array(labels).astype(dtype=np.uint8)

    #convert the int numpy array into a one-hot matrix
  #  labels_onehot = (np.arange(NUM_LABELS) == labels_np[:, None]).astype(np.float32))
#