from openpyxl import utils
from openpyxl import load_workbook
import csv
import os


def FindingXL(directory):
    ofile = open('rtc_vehicles.csv', "wb")
    writer = csv.writer(ofile, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
    header = []
    header.append('Sheet Name')
    header.append('Vehicle')
    header.append('Date')
    header.append('KM Covered')
    header.append('Amount$')
    writer.writerow(header)

    for name in os.listdir(directory): # lists directories in an arbitrary order i.e
        if name.endswith(".xlsx"): # let's look for xlsx files
            wb = load_workbook(directory+'/'+name, data_only=True)
            Crawler(wb,writer,name)

    ofile.close()

def Crawler(wb,writer,name):
    print name
    startRow = 0
    endRow = 0
    startCol = 0
    endCol = 0
    for sheet in wb.worksheets:
        title = sheet.title
        bEndRowFound = False
        for row in sheet.iter_rows():
            for cell in row:
                cellValue = cell.value
                if cellValue == 'SL #':
                    coord = cell.coordinate
                    t = utils.coordinate_from_string(coord)
                    startCol = utils.column_index_from_string(t[0])-1
                    startRow = t[1]+3

                if type(cellValue) is unicode:
                    if cellValue.strip() == 'Total':
                        bEndRowFound = True
                        coord = cell.coordinate
                        t = utils.coordinate_from_string(coord)
                        endCol = utils.column_index_from_string(t[0])+8
                        endRow = t[1]-1
            if(bEndRowFound):
                break

        for irow in range(startRow, endRow):
            print name
            row = sheet[irow]
            rowOut = []
            rowOut.append(title)
            rowOut.append(name)
            for icell in range(startCol, endCol):
                cell = row[icell]
                cellValue = cell.value
                coord = cell.coordinate
                t = utils.coordinate_from_string(coord)
                cellColumn = t[0]
                cellColumNumber = utils.column_index_from_string(cellColumn)
                if cellColumNumber==startCol+2:
                    rowOut.append(cellValue)
                    print coord
                    print cellValue
                elif cellColumNumber==startCol+5:
                    rowOut.append(cellValue)
                    print coord
                    print cellValue
                elif cellColumNumber==startCol+10:
                    rowOut.append(cellValue)
                    print coord
                    print cellValue

            writer.writerow(rowOut)

defaultDirectory = "/Users/pc/Desktop/DataScientists/Admin"
FindingXL(defaultDirectory)